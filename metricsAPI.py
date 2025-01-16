import requests
import json
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import pandas as pd
from io import BytesIO
import matplotlib.pyplot as plt
from datetime import datetime
from tkinter import Tk, filedialog
import os

# Configure logging with dynamic filename
log_filename = "metrics_debug.log"
if os.path.exists(log_filename):
    log_filename = f"metrics_debug_{datetime.now().strftime('%Y%m%d%H%M%S')}.log"
logging.basicConfig(filename=log_filename, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Define thresholds for green, yellow, red
thresholds = {
    "Processor": {"green": 50, "yellow": 90, "red": 100},
    "Memory": {"green": 30, "yellow": 95, "red": 100},
    "Average Disk Used Percentage": {"green": 60, "yellow": 85, "red": 100},
    "Average Disk Idletime Percentage": {"green": 60, "yellow": 85, "red": 100},
    "Disk Transfer Per Second": {"green": 60, "yellow": 85, "red": 100},
    "Average Disk Queue Length": {"green": 60, "yellow": 85, "red": 100},
    "Network Adapter In": {"green": 20, "yellow": 70, "red": 100},
    "Network Adapter Out": {"green": 20, "yellow": 70, "red": 100}
}

def fetch_metrics(api_url, headers, metric, entity_filter, management_zone, start_time):
    """
    Fetches metrics from Dynatrace using the Metrics API.
    """
    url = f"{api_url}?metricSelector={metric}&from={start_time}&entitySelector={entity_filter}&mzSelector=mzName(\"{management_zone}\")"
    logging.debug(f"Fetching data from URL: {url}")
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    logging.debug(f"Response for {metric}: {response.json()}")
    return response.json()

def create_chart(chart_data, title, metric_name):
    """
    Create a bar chart for a given metric and save it to a BytesIO stream.
    Apply color coding based on thresholds.
    """
    logging.debug(f"Creating chart for {metric_name} with data: {chart_data}")
    colors = []
    for value in chart_data[metric_name]:
        if value <= thresholds[metric_name]["green"]:
            colors.append("green")
        elif value <= thresholds[metric_name]["yellow"]:
            colors.append("yellow")
        else:
            colors.append("red")

    plt.figure(figsize=(8, 4))
    plt.bar(chart_data['Time'], chart_data[metric_name], color=colors)
    plt.title(title)
    plt.xlabel('Time')
    plt.ylabel(metric_name)
    plt.xticks(rotation=45, ha='right')
    chart_stream = BytesIO()
    plt.tight_layout()
    plt.savefig(chart_stream, format='png')
    plt.close()
    chart_stream.seek(0)
    return chart_stream

def add_title_block(sheet, management_zone, start_time, num_servers):
    """
    Add a title block to the top of the first sheet with report details.
    """
    report_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    duration = "Weekly" if "1w" in start_time else "Daily" if "1d" in start_time else "Custom"
    title_content = [
        f"Team Name/Management Zone: {management_zone}",
        f"Report Time: {report_time}",
        f"Report Duration: {start_time}",
        f"Number of Servers: {num_servers}"
    ]
    for idx, line in enumerate(title_content, start=1):
        sheet.cell(row=idx, column=1, value=line)
        sheet.cell(row=idx, column=1).font = Font(bold=True)

def aggregate_metric_data(metric_data, metric_name):
    """
    Aggregate timestamps and values from metric_data.
    """
    aggregated = {}
    for entity in metric_data.get("entities", []):
        host = entity.get("displayName", "Unknown Host")
        timestamps = entity.get("dataPoints", {}).get("timestamps", [])
        values = entity.get("dataPoints", {}).get("values", [])
        if not timestamps or not values:
            logging.warning(f"No data points found for metric: {metric_name}, host: {host}")
            continue
        if host not in aggregated:
            aggregated[host] = {}
        aggregated[host][metric_name] = list(zip(timestamps, values))
    return aggregated

def generate_excel_report(aggregated_data, management_zone, start_time, output_filename):
    """
    Generate a new Excel report with a title block and charts.
    """
    if os.path.exists(output_filename):
        output_filename = f"Aggregated_Dynatrace_Report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    workbook = Workbook()
    title_sheet = workbook.active
    title_sheet.title = "Report Summary"
    num_servers = len(aggregated_data)
    add_title_block(title_sheet, management_zone, start_time, num_servers)

    for host, metrics_data in aggregated_data.items():
        logging.debug(f"Generating sheet for host: {host} with metrics: {metrics_data}")
        if not metrics_data:
            logging.warning(f"No metrics data found for host: {host}. Skipping sheet creation.")
            continue

        sheet = workbook.create_sheet(title=host[:31])
        sheet.cell(row=1, column=1, value="Metric")
        sheet.cell(row=1, column=2, value="Time")
        sheet.cell(row=1, column=3, value="Value")

        row_idx = 2
        for metric_name, data_points in metrics_data.items():
            logging.debug(f"Processing metric: {metric_name} with data points: {data_points}")
            if not data_points:
                logging.warning(f"No data points for metric: {metric_name} on host: {host}. Skipping.")
                continue

            for time, value in data_points:
                sheet.cell(row=row_idx, column=1, value=metric_name)
                sheet.cell(row=row_idx, column=2, value=time)
                sheet.cell(row=row_idx, column=3, value=value)
                row_idx += 1

            chart_stream = create_chart(pd.DataFrame(data_points, columns=['Time', metric_name]), f"{metric_name} Trend", metric_name)
            img = Image(chart_stream)
            sheet.add_image(img, f"E{row_idx - len(data_points)}")

    workbook.save(output_filename)
    logging.info(f"Excel report saved to {output_filename}")
    with open("aggregated_data.json", "w") as f:
        if os.path.exists("aggregated_data.json"):
            f = open(f"aggregated_data_{datetime.now().strftime('%Y%m%d%H%M%S')}.json", "w")
        json.dump(aggregated_data, f, indent=4)
        logging.info("Aggregated data saved to aggregated_data.json for review.")

def main():
    Tk().withdraw()
    print("Enter Dynatrace API Details:")
    api_url = input("Enter the Dynatrace Metrics API URL: ").strip()
    api_token = input("Enter your API Token: ").strip()
    management_zone = input("Enter the Management Zone: ").strip()
    start_time = input("Enter the start time (e.g., now-1w): ").strip()

    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Accept": "application/json; charset=utf-8"
    }

    metrics = {
        "Processor": "builtin:host.cpu.usage",
        "Memory": "builtin:host.mem.usage",
        "Average Disk Used Percentage": "builtin:host.disk.usedPct",
        "Average Disk Idletime Percentage": "com.dynatrace.extension.host-observability.disk.usage.idle.percent",
        "Disk Transfer Per Second": "com.dynatrace.extension.host-observability.disk.transfer.persec",
        "Average Disk Queue Length": "builtin:host.disk.queueLength",
        "Network Adapter In": "builtin:host.net.nic.trafficIn",
        "Network Adapter Out": "builtin:host.net.nic.trafficOut"
    }

    aggregated_data = {}
    for metric_name, metric_selector in metrics.items():
        logging.info(f"Fetching data for {metric_name}...")
        metric_data = fetch_metrics(api_url, headers, metric_selector, 'type("HOST")', management_zone, start_time)
        logging.debug(f"Fetched metric data for {metric_name}: {metric_data}")

        # Aggregate metric data
        metric_aggregated = aggregate_metric_data(metric_data, metric_name)
        for host, metrics in metric_aggregated.items():
            if host not in aggregated_data:
                aggregated_data[host] = {}
            aggregated_data[host].update(metrics)

    logging.debug(f"Final aggregated data: {aggregated_data}")
    output_filename = "Aggregated_Dynatrace_Report.xlsx"
    print("Generating the Excel report...")
    generate_excel_report(aggregated_data, management_zone, start_time, output_filename)

if __name__ == "__main__":
    main()
