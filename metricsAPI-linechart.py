import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import matplotlib.pyplot as plt
from datetime import datetime
from tkinter import Tk, filedialog

# Define thresholds for green, yellow, red
thresholds = {
    "Processor": {"green": 50, "yellow": 90, "red": 100},
    "Memory": {"green": 30, "yellow": 95, "red": 100},
    "Average Disk Used Percentage": {"green": 60, "yellow": 85, "red": 100},
    "Average Disk Utilzation Time": {"green": 60, "yellow": 85, "red": 100},
    "Disk Write Time Per Second": {"green": 60, "yellow": 900, "red": 1000},
    "Average Disk Queue Length": {"green": 75, "yellow": 200, "red": 500},
    "Network Adapter In": {"green": 500000000, "yellow": 1000000000, "red": 1900000000},
    "Network Adapter Out": {"green": 500000000, "yellow": 2000000000, "red": 2500000000}
}

def fetch_metrics(api_url, headers, metric, entity_filter, management_zone, start_time):
    """
    Fetches metrics from Dynatrace using the Metrics API.
    """
    url = f"{api_url}?metricSelector={metric}&from={start_time}&entitySelector={entity_filter}&mzSelector=mzName({management_zone})"
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()

def create_chart(chart_data, title, metric_name):
    """
    Create a line chart for a given metric and save it to a BytesIO stream.
    """
    plt.figure(figsize=(8, 4))
    plt.plot(chart_data['Time'], chart_data[metric_name], marker='o', linestyle='-', color='blue', label=metric_name)
    plt.title(title)
    plt.xlabel('Time')
    plt.ylabel(metric_name)
    plt.xticks(rotation=45, ha='right')
    plt.grid(True)
    plt.legend()

    chart_stream = BytesIO()
    plt.tight_layout()
    plt.savefig(chart_stream, format='png')
    plt.close()

    chart_stream.seek(0)
    return chart_stream

def add_title_block(sheet, management_zone, start_time, metrics, num_servers):
    """
    Add a title block to the top of the first sheet with report details.
    """
    report_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    duration = "Weekly" if "1w" in start_time else "Daily" if "1d" in start_time else "Custom"

    title_content = [
        f"Team Name/Management Zone: {management_zone}",
        f"Report Time: {report_time}",
        f"Report Duration: {start_time}",
        f"Number of Servers: {num_servers}",
        f"Resources: {', '.join(metrics)}",
    ]

    for idx, line in enumerate(title_content, start=1):
        sheet.cell(row=idx, column=1, value=line)
        sheet.cell(row=idx, column=1).font = Font(bold=True)

def generate_excel_report(aggregated_data, management_zone, start_time, metrics, output_filename):
    """
    Generate a new Excel report with a title block and line charts.
    """
    workbook = Workbook()
    title_sheet = workbook.active
    title_sheet.title = "Report Summary"

    num_servers = len(aggregated_data)
    add_title_block(title_sheet, management_zone, start_time, metrics, num_servers)

    for dimension, df in aggregated_data.items():
        sheet = workbook.create_sheet(title=str(dimension)[:31])

        for r_idx, row in enumerate(df.itertuples(index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        for metric_name in df.columns:
            if metric_name not in ['Dimension', 'Time']:
                chart_stream = create_chart(df, f"{metric_name} Trend for {dimension}", metric_name)

                img = Image(chart_stream)
                sheet.add_image(img, f"B{r_idx + 2}")

    workbook.save(output_filename)
    print(f"Excel report saved to {output_filename}")

def main():
    Tk().withdraw()
    print("Please select the existing Dynatrace report file (Excel).")
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        print("No file selected. Exiting...")
        return

    management_zone = input("Enter the Management Zone: ").strip()
    start_time = "now-1w"
    metrics = [
        "Processor", "Memory", "Average Disk Used Percentage",
        "Average Disk Utilzation Time", "Disk Write Time Per Second",
        "Average Disk Queue Length", "Network Adapter In", "Network Adapter Out"
    ]

    print("Aggregating data from the existing report...")
    aggregated_data = aggregate_data_from_existing_report(file_path)

    output_filename = f"{management_zone.replace(':', '').replace(' ', '_')}-Aggregated_Dynatrace_Report-{datetime.now().strftime('%Y%m%d')}.xlsx"
    print("Generating the Excel report...")
    generate_excel_report(aggregated_data, management_zone, start_time, metrics, output_filename)
    print(f"Report saved to {output_filename}")

if __name__ == "__main__":
    main()
